import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


def generate_excel(expenses: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    headers = ["Fecha", "Descripcion", "Monto", "Categoria", "Fuente"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, exp in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=exp["date"]).border = thin_border
        ws.cell(row=row, column=2, value=exp["description"]).border = thin_border
        monto_cell = ws.cell(row=row, column=3, value=exp["amount"])
        monto_cell.number_format = '#,##0.00'
        monto_cell.border = thin_border
        ws.cell(row=row, column=4, value=exp.get("category", "Otros")).border = thin_border
        ws.cell(row=row, column=5, value=exp.get("source", "")).border = thin_border

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 25

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
